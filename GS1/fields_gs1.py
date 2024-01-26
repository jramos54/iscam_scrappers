import pandas as pd
import openpyxl,json,os
import numpy as np

def extraer_encabezados_excel(archivo_excel):
    try:
        # Abrir el archivo de Excel con openpyxl
        workbook = openpyxl.load_workbook(archivo_excel, read_only=True)
        
        encabezados_por_pestaña = {}
        
        # Nombres de pestañas específicos que deseas procesar
        pestañas_deseadas = ["Relationships", "Model", "Entities"]
        
        # Iterar a través de todas las pestañas
        for nombre_pestaña in workbook.sheetnames:
            if nombre_pestaña in pestañas_deseadas:
                sheet = workbook[nombre_pestaña]
                
                # Obtener los encabezados de la segunda fila
                encabezados = []
                for row in sheet.iter_rows(min_row=2, max_row=2, values_only=True):
                    encabezados.extend(row)
                
                # Agregar los encabezados al diccionario de la pestaña
                encabezados_por_pestaña[nombre_pestaña] = encabezados
        
        # Guardar la salida en un archivo JSON
        with open('encabezados.json', 'w', encoding='utf-8') as json_file:
            json.dump(encabezados_por_pestaña, json_file, ensure_ascii=False, indent=4)
        
        return "Los encabezados se han guardado en 'encabezados.json'"
    except Exception as e:
        return str(e)

def count_non_empty_cells(file_path, sheets):
    """
    Cuenta las celdas no vacías en cada columna de las hojas especificadas de un archivo Excel.

    Parámetros:
    file_path (str): La ruta del archivo Excel.
    sheets (list): Lista de nombres de las hojas que se desean importar.

    Retorna:
    dict: Un diccionario donde cada clave es el nombre de la hoja y el valor es otro diccionario
          con el conteo de celdas no vacías para cada columna.
    """
    non_empty_counts = {}
    # for sheet in sheets:
    #     # Importar la hoja, omitiendo la primera fila (header=1)
    #     df = pd.read_excel(file_path, sheet_name=sheet, header=1)
    #     # Contar celdas no vacías para cada columna
    #     counts = df.count().to_dict()
    #     non_empty_counts[sheet] = counts
    for sheet in sheets:
        # Importar la hoja, omitiendo la primera fila (header=1)
        df = pd.read_excel(file_path, sheet_name=sheet, header=1)
        # Contar celdas no vacías para cada columna y excluir columnas con conteo cero
        counts = {col: count for col, count in df.count().items() if count > 0}
        non_empty_counts[sheet] = counts
    
    with open('cuantificacion.json', 'w', encoding='utf-8') as json_file:
        json.dump(non_empty_counts, json_file, ensure_ascii=False, indent=4)
                
    return non_empty_counts

def excel_to_json_null(input_excel_file, output_json_file, sheet_names):
    try:
        # Lee el archivo Excel
        xls = pd.ExcelFile(input_excel_file)
        data = {}

        for sheet_name in sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=1)  # Comienza desde la fila 2

            # Elimina las filas y columnas con todos los valores en blanco (NaN, null, None)
            df.dropna(axis=0, how='all', inplace=True)
            df.dropna(axis=1, how='all', inplace=True)

            # Convierte la hoja de datos a una lista de diccionarios
            # excluyendo explícitamente las claves con valores NaN, null o None
            sheet_data = []
            for _, row in df.iterrows():
                row_dict = {col: row[col] for col in df.columns if pd.notna(row[col])}
                sheet_data.append(row_dict)

            # Almacena la lista de diccionarios en el diccionario principal
            data[sheet_name] = sheet_data

        # Convierte el diccionario a formato JSON
        with open(output_json_file, 'w') as json_file:
            json.dump(data, json_file, default=str, indent=4)

        print(f"El archivo JSON '{output_json_file}' ha sido creado con éxito.")
    
    except Exception as e:
        print(f"Se produjo un error al procesar el archivo Excel: {str(e)}")

def excel_to_json(input_excel_file, output_json_file, sheet_names):
    try:
        # Lee el archivo Excel
        xls = pd.ExcelFile(input_excel_file)
        data = {}

        for sheet_name in sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=1)  # Comienza desde la fila 2

            # Elimina las filas con todos los valores en blanco
            df.dropna(axis=0, how='all', inplace=True)

            # Reemplaza los valores NaN, null o None con None
            df = df.where(pd.notnull(df), None)

            # Convierte la hoja de datos a un diccionario
            # Aquí, las celdas con None serán ignoradas en el diccionario
            sheet_dict = df.to_dict(orient='records')

            # Almacena el diccionario en el diccionario principal usando el nombre de la hoja como clave
            data[sheet_name] = sheet_dict

        # Convierte el diccionario a formato JSON
        with open(output_json_file, 'w') as json_file:
            json.dump(data, json_file, default=str, indent=4)  # Usamos default=str para manejar objetos no serializables

        print(f"El archivo JSON '{output_json_file}' ha sido creado con éxito.")
    
    except Exception as e:
        print(f"Se produjo un error al procesar el archivo Excel: {str(e)}")

def excel_to_dict(input_excel_file, sheet_names):
    try:
        xls = pd.ExcelFile(input_excel_file)
        data = {}

        for sheet_name in sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=1)

            df.dropna(axis=0, how='all', inplace=True)
            df.dropna(axis=1, how='all', inplace=True)

            sheet_data = []
            for _, row in df.iterrows():
                row_dict = {col: row[col].strftime('%Y-%m-%d %H:%M:%S') if isinstance(row[col], pd.Timestamp) else row[col] for col in df.columns if pd.notna(row[col])}
                sheet_data.append(row_dict)

            data[sheet_name] = sheet_data

        return data

    except Exception as e:
        print(f"Se produjo un error al procesar el archivo Excel: {str(e)}")
        return None
    
def save_dict_to_json(data, output_json_file):
    try:
        with open(output_json_file, 'w', encoding='utf-8') as json_file:  # Añadir encoding='utf-8'
            json.dump(data, json_file, ensure_ascii=False, default=str, indent=4)  # Añadir ensure_ascii=False
        print(f"El diccionario ha sido guardado en el archivo JSON '{output_json_file}' con éxito.")
    except Exception as e:
        print(f"Se produjo un error al guardar el diccionario como JSON: {str(e)}")

def list_files_in_directory(directory):
    file_list = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            relative_path = os.path.relpath(os.path.join(root, file), directory)
            file_list.append(relative_path.replace('\\', '/'))
    return file_list

def extract_values_from_dicts(dict_list, key_list):
    result = {}

    for key_dict in key_list:
        for friendly_key, real_key in key_dict.items():  # Obtiene la clave amigable y la clave real
            unique_values = set()

            for dictionary in dict_list:
                if real_key in dictionary:
                    unique_values.add(dictionary[real_key])

            result[friendly_key] = list(unique_values)

    return result

if __name__=="__main__":
    
    directory = r'excel_files'
    items_json="items_examples.json"
    keys_search="key_search.json"
    
    sheets_to_import = ["Relationships", "Model", "Entities"]
    keys=[{"nombre_generico":"Name"},{"marca":"Brand Name BMSid 3541"},{"publicador":"Information provider party name BMSid 85"}]
    
    all_data = {}  # Diccionario para acumular todos los datos de los archivos

    archivos_excel = list_files_in_directory(directory)
    
    for archivo in archivos_excel:
        data = excel_to_dict(os.path.join(directory, archivo), sheets_to_import)
        if data:
            # Combina los datos del archivo actual con los datos acumulados
            all_data.update(data)

    # print(json.dumps(all_data, indent=4))
    print(len(all_data['Entities']))
    
    save_dict_to_json(data, items_json)
    
    keys_result=extract_values_from_dicts(all_data['Entities'],keys)
    
    print(len(keys_result["nombre_generico"]))
    print(len(keys_result["marca"]))
    print(len(keys_result["publicador"]))
    
    save_dict_to_json(keys_result, keys_search)

    # resultado = extraer_encabezados_excel(archivo_excel)
    # non_empty_cells_count = count_non_empty_cells(archivo_excel, sheets_to_import)
    # print(json.dumps(non_empty_cells_count,indent=4))
    # excel_to_json(archivo_excel, 'ejemplo.json', sheets_to_import)
    # excel_to_json_null(archivo_excel, 'ejemplo_null.json', sheets_to_import)


